from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
import pyautogui
pyautogui.PAUSE = 2
pyautogui.FAILSAFE = True
import os

wb = openpyxl.load_workbook("C:/Users/silambarasan/ZF Friedrichshafen AG/Global ISC & Business Improvement Team - Documents/Python Projects/eDOC_download/datasheet.xlsx")
sh = wb['Sheet1']
rows = sh.max_row
cols = sh.max_column


browser = webdriver.Edge("C:/Users/silambarasan/ZF Friedrichshafen AG/Global ISC & Business Improvement Team - Documents/Python Projects/msedgedriver.exe")

browser.get('edge://settings/downloads')
browser.maximize_window()
time.sleep(3)
pyautogui.PAUSE = 2  # Setup of pyautogui parameters
pyautogui.FAILSAFE = True  # this enables a FAILSAFE operation:
screenWidth, screenHeight = pyautogui.size()
pcng = pyautogui.locateCenterOnScreen('pics/pchange.png')
pyautogui.click(pcng)
pyautogui.write('docs')
pyautogui.press('enter', presses=2)
browser.get('https://app11.jaggaer.com/modules/myp4t/#_menu/relationship_contract_management')
username = browser.find_element_by_id('username')
username.send_keys('GBSRPA_user')
password = browser.find_element_by_id('tpass')
password.send_keys('April@2021')
browser.maximize_window()
cred_submit = browser.find_element_by_class_name('fieldButton')
cred_submit.click()

for r in range(2,rows+1):
    for c in range(1, cols):
                    time.sleep(10)
                    screenWidth, screenHeight = pyautogui.size()
                    search = pyautogui.locateCenterOnScreen('pics/search.png')
                    pyautogui.click(search)
                    time.sleep(2)
                    c_id = pyautogui.locateCenterOnScreen('pics/contract.png')
                    pyautogui.click(c_id)
                    pyautogui.press('tab', presses=1)
                    con_id = (sh.cell(row=r, column=1).value)
                    pyautogui.typewrite(str(con_id))
                    pyautogui.press('enter', presses=1)
                    time.sleep(5)
                    opn = pyautogui.locateCenterOnScreen('pics/open.png')
                    pyautogui.click(opn)
                    time.sleep(3)
                    pyautogui.press('pagedown', presses=1)
                    cpy = pyautogui.locateCenterOnScreen('pics/copy.png')
                    pyautogui.click(cpy)
                    pyautogui.press('Esc', presses=1)
                    pyautogui.press('pageup', presses=1)
                    bck = pyautogui.locateCenterOnScreen('pics/back.png')
                    pyautogui.click(bck)
                    sh.cell(row=r, column=4).value = "Downloaded"
                    wb.save("C:/Users/silambarasan/ZF Friedrichshafen AG/Global ISC & Business Improvement Team - Documents/Python Projects/eDOC_download/datasheet.xlsx")
                    wb.close()
                    break